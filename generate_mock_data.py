import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

def generate_data():
    print("Generating raw supply chain data...")
    os.makedirs("data", exist_ok=True)
    os.makedirs("visualizations", exist_ok=True)
    
    # 1. Nodes Data
    nodes = [
        # Tier 3 - Mining
        {"Node_ID": "MIN_LITH_CHL", "Name": "Lithium Mine - Atacama", "Tier": "Tier-3 (Mining)", "Location": "Atacama Desert", "Country": "Chile", "Latitude": -23.5000, "Longitude": -68.3000, "Base_Capacity_Tons": 15000, "Cost_Per_Ton": 8000, "Scenario_Scope": "Both", "FEOC_Status": "No"},
        {"Node_ID": "MIN_LITH_AUS", "Name": "Lithium Mine - Greenbushes", "Tier": "Tier-3 (Mining)", "Location": "Greenbushes", "Country": "Australia", "Latitude": -33.8500, "Longitude": 116.0500, "Base_Capacity_Tons": 12000, "Cost_Per_Ton": 9500, "Scenario_Scope": "Both", "FEOC_Status": "No"},
        {"Node_ID": "MIN_LITH_CAN", "Name": "Lithium Mine - Quebec", "Tier": "Tier-3 (Mining)", "Location": "La Corne", "Country": "Canada", "Latitude": 48.3300, "Longitude": -77.9800, "Base_Capacity_Tons": 8000, "Cost_Per_Ton": 11000, "Scenario_Scope": "To-Be", "FEOC_Status": "No"},
        {"Node_ID": "MIN_COB_DRC", "Name": "Cobalt Mine - Kolwezi", "Tier": "Tier-3 (Mining)", "Location": "Lualaba Province", "Country": "DRC", "Latitude": -10.7100, "Longitude": 25.4700, "Base_Capacity_Tons": 10000, "Cost_Per_Ton": 25000, "Scenario_Scope": "Both", "FEOC_Status": "No"},
        {"Node_ID": "MIN_COB_CAN", "Name": "Cobalt Mine - Sudbury", "Tier": "Tier-3 (Mining)", "Location": "Sudbury", "Country": "Canada", "Latitude": 46.4900, "Longitude": -81.0100, "Base_Capacity_Tons": 3000, "Cost_Per_Ton": 32000, "Scenario_Scope": "To-Be", "FEOC_Status": "No"},
        {"Node_ID": "MIN_NIC_IDN", "Name": "Nickel Mine - Sorowako", "Tier": "Tier-3 (Mining)", "Location": "Sulawesi", "Country": "Indonesia", "Latitude": -2.5300, "Longitude": 121.3600, "Base_Capacity_Tons": 20000, "Cost_Per_Ton": 12000, "Scenario_Scope": "Both", "FEOC_Status": "Yes"},
        {"Node_ID": "MIN_NIC_CAN", "Name": "Nickel Mine - Voisey Bay", "Tier": "Tier-3 (Mining)", "Location": "Labrador", "Country": "Canada", "Latitude": 56.3100, "Longitude": -62.1000, "Base_Capacity_Tons": 10000, "Cost_Per_Ton": 15000, "Scenario_Scope": "To-Be", "FEOC_Status": "No"},
        {"Node_ID": "MIN_GRA_CHN", "Name": "Graphite Mine - Inner Mongolia", "Tier": "Tier-3 (Mining)", "Location": "Inner Mongolia", "Country": "China", "Latitude": 40.8100, "Longitude": 111.6900, "Base_Capacity_Tons": 18000, "Cost_Per_Ton": 3000, "Scenario_Scope": "As-Is", "FEOC_Status": "Yes"},
        {"Node_ID": "MIN_GRA_USA", "Name": "Synthetic Graphite - Ohio", "Tier": "Tier-3 (Mining)", "Location": "Cleveland", "Country": "USA", "Latitude": 41.4900, "Longitude": -81.6900, "Base_Capacity_Tons": 9000, "Cost_Per_Ton": 7500, "Scenario_Scope": "To-Be", "FEOC_Status": "No"},
        
        # Tier 2 - Refining & Processing
        {"Node_ID": "REF_LITH_CHL", "Name": "Lithium Carbonate - Antofagasta", "Tier": "Tier-2 (Refining)", "Location": "Antofagasta", "Country": "Chile", "Latitude": -23.6500, "Longitude": -70.4000, "Base_Capacity_Tons": 14000, "Cost_Per_Ton": 2000, "Scenario_Scope": "Both", "FEOC_Status": "No"},
        {"Node_ID": "REF_COB_CHN", "Name": "Cobalt Refinery - Jinchuan", "Tier": "Tier-2 (Refining)", "Location": "Jinchang", "Country": "China", "Latitude": 38.5000, "Longitude": 102.1900, "Base_Capacity_Tons": 9000, "Cost_Per_Ton": 4000, "Scenario_Scope": "As-Is", "FEOC_Status": "Yes"},
        {"Node_ID": "REF_COB_FIN", "Name": "Cobalt Refinery - Kokkola", "Tier": "Tier-2 (Refining)", "Location": "Kokkola", "Country": "Finland", "Latitude": 63.8300, "Longitude": 23.1200, "Base_Capacity_Tons": 4000, "Cost_Per_Ton": 6000, "Scenario_Scope": "To-Be", "FEOC_Status": "No"},
        {"Node_ID": "REF_NIC_CHN", "Name": "Nickel Smelter - Jinchuan", "Tier": "Tier-2 (Refining)", "Location": "Jinchang", "Country": "China", "Latitude": 38.5000, "Longitude": 102.1900, "Base_Capacity_Tons": 18000, "Cost_Per_Ton": 3000, "Scenario_Scope": "As-Is", "FEOC_Status": "Yes"},
        {"Node_ID": "REF_NIC_CAN", "Name": "Nickel Refinery - Fort Sask", "Tier": "Tier-2 (Refining)", "Location": "Fort Saskatchewan", "Country": "Canada", "Latitude": 53.7100, "Longitude": -113.2100, "Base_Capacity_Tons": 9000, "Cost_Per_Ton": 4500, "Scenario_Scope": "To-Be", "FEOC_Status": "No"},
        {"Node_ID": "REF_GRA_CHN", "Name": "Anode Material Plant - Heilongjiang", "Tier": "Tier-2 (Refining)", "Location": "Harbin", "Country": "China", "Latitude": 45.7500, "Longitude": 126.6300, "Base_Capacity_Tons": 16000, "Cost_Per_Ton": 1500, "Scenario_Scope": "As-Is", "FEOC_Status": "Yes"},
        {"Node_ID": "REF_GRA_USA", "Name": "Anode Material Plant - Tennessee", "Tier": "Tier-2 (Refining)", "Location": "Chattanooga", "Country": "USA", "Latitude": 35.0400, "Longitude": -85.3000, "Base_Capacity_Tons": 8000, "Cost_Per_Ton": 3000, "Scenario_Scope": "To-Be", "FEOC_Status": "No"},
        
        # Tier 1.5 - Cathode Active Material (CAM) Synthesis
        {"Node_ID": "CAM_KOR", "Name": "CAM Plant - Busan", "Tier": "Tier-1.5 (CAM)", "Location": "Busan", "Country": "South Korea", "Latitude": 35.1000, "Longitude": 129.0400, "Base_Capacity_Tons": 25000, "Cost_Per_Ton": 15000, "Scenario_Scope": "Both", "FEOC_Status": "No"},
        {"Node_ID": "CAM_USA", "Name": "CAM Plant - Indiana", "Tier": "Tier-1.5 (CAM)", "Location": "Kokomo", "Country": "USA", "Latitude": 40.4800, "Longitude": -86.1300, "Base_Capacity_Tons": 15000, "Cost_Per_Ton": 19000, "Scenario_Scope": "To-Be", "FEOC_Status": "No"},
        
        # Tier 1 - Cell Assembly (Gigafactory)
        {"Node_ID": "GIGA_USA", "Name": "Michigan Gigafactory", "Tier": "Tier-1 (Assembly)", "Location": "Detroit", "Country": "USA", "Latitude": 42.3300, "Longitude": -83.0400, "Base_Capacity_Tons": 35000, "Cost_Per_Ton": 10000, "Scenario_Scope": "Both", "FEOC_Status": "No"}
    ]
    
    # 2. Routes Data
    routes = [
        # As-Is Scenario Routes
        {"Route_ID": "R_ASIS_01", "Origin": "MIN_LITH_CHL", "Destination": "REF_LITH_CHL", "Mode": "Truck", "Lead_Time_Days": 1, "Freight_Cost_Per_Ton": 200, "Scenario": "As-Is"},
        {"Route_ID": "R_ASIS_02", "Origin": "REF_LITH_CHL", "Destination": "CAM_KOR", "Mode": "Maritime", "Lead_Time_Days": 28, "Freight_Cost_Per_Ton": 1200, "Scenario": "As-Is"},
        {"Route_ID": "R_ASIS_03", "Origin": "MIN_LITH_AUS", "Destination": "CAM_KOR", "Mode": "Maritime", "Lead_Time_Days": 18, "Freight_Cost_Per_Ton": 900, "Scenario": "As-Is"},
        {"Route_ID": "R_ASIS_04", "Origin": "MIN_COB_DRC", "Destination": "REF_COB_CHN", "Mode": "Maritime", "Lead_Time_Days": 32, "Freight_Cost_Per_Ton": 1500, "Scenario": "As-Is"},
        {"Route_ID": "R_ASIS_05", "Origin": "REF_COB_CHN", "Destination": "CAM_KOR", "Mode": "Maritime", "Lead_Time_Days": 3, "Freight_Cost_Per_Ton": 300, "Scenario": "As-Is"},
        {"Route_ID": "R_ASIS_06", "Origin": "MIN_NIC_IDN", "Destination": "REF_NIC_CHN", "Mode": "Maritime", "Lead_Time_Days": 12, "Freight_Cost_Per_Ton": 600, "Scenario": "As-Is"},
        {"Route_ID": "R_ASIS_07", "Origin": "REF_NIC_CHN", "Destination": "CAM_KOR", "Mode": "Maritime", "Lead_Time_Days": 3, "Freight_Cost_Per_Ton": 300, "Scenario": "As-Is"},
        {"Route_ID": "R_ASIS_08", "Origin": "MIN_GRA_CHN", "Destination": "REF_GRA_CHN", "Mode": "Rail", "Lead_Time_Days": 4, "Freight_Cost_Per_Ton": 400, "Scenario": "As-Is"},
        {"Route_ID": "R_ASIS_09", "Origin": "REF_GRA_CHN", "Destination": "CAM_KOR", "Mode": "Maritime", "Lead_Time_Days": 4, "Freight_Cost_Per_Ton": 350, "Scenario": "As-Is"},
        {"Route_ID": "R_ASIS_10", "Origin": "CAM_KOR", "Destination": "GIGA_USA", "Mode": "Maritime + Rail (LA Port)", "Lead_Time_Days": 25, "Freight_Cost_Per_Ton": 1800, "Scenario": "As-Is"},
        
        # To-Be Scenario Routes (Resilient, Domestic/FTA, FEOC-Free)
        {"Route_ID": "R_TOBE_01", "Origin": "MIN_LITH_CHL", "Destination": "REF_LITH_CHL", "Mode": "Truck", "Lead_Time_Days": 1, "Freight_Cost_Per_Ton": 200, "Scenario": "To-Be"},
        {"Route_ID": "R_TOBE_02", "Origin": "REF_LITH_CHL", "Destination": "CAM_USA", "Mode": "Maritime (East Coast)", "Lead_Time_Days": 22, "Freight_Cost_Per_Ton": 1400, "Scenario": "To-Be"},
        {"Route_ID": "R_TOBE_03", "Origin": "MIN_LITH_CAN", "Destination": "CAM_USA", "Mode": "Rail", "Lead_Time_Days": 6, "Freight_Cost_Per_Ton": 800, "Scenario": "To-Be"},
        {"Route_ID": "R_TOBE_04", "Origin": "MIN_COB_CAN", "Destination": "REF_COB_FIN", "Mode": "Maritime (Transatlantic)", "Lead_Time_Days": 16, "Freight_Cost_Per_Ton": 1100, "Scenario": "To-Be"},
        {"Route_ID": "R_TOBE_05", "Origin": "REF_COB_FIN", "Destination": "CAM_USA", "Mode": "Maritime (Transatlantic)", "Lead_Time_Days": 14, "Freight_Cost_Per_Ton": 1300, "Scenario": "To-Be"},
        {"Route_ID": "R_TOBE_06", "Origin": "MIN_NIC_CAN", "Destination": "REF_NIC_CAN", "Mode": "Rail", "Lead_Time_Days": 5, "Freight_Cost_Per_Ton": 600, "Scenario": "To-Be"},
        {"Route_ID": "R_TOBE_07", "Origin": "REF_NIC_CAN", "Destination": "CAM_USA", "Mode": "Rail", "Lead_Time_Days": 4, "Freight_Cost_Per_Ton": 500, "Scenario": "To-Be"},
        {"Route_ID": "R_TOBE_08", "Origin": "MIN_GRA_USA", "Destination": "REF_GRA_USA", "Mode": "Truck", "Lead_Time_Days": 2, "Freight_Cost_Per_Ton": 300, "Scenario": "To-Be"},
        {"Route_ID": "R_TOBE_09", "Origin": "REF_GRA_USA", "Destination": "CAM_USA", "Mode": "Truck", "Lead_Time_Days": 2, "Freight_Cost_Per_Ton": 250, "Scenario": "To-Be"},
        {"Route_ID": "R_TOBE_10", "Origin": "CAM_USA", "Destination": "GIGA_USA", "Mode": "Truck/Rail", "Lead_Time_Days": 2, "Freight_Cost_Per_Ton": 400, "Scenario": "To-Be"}
    ]
    
    # 3. Risk Registry for 2026
    risk_library = [
        {
            "Risk_ID": "RSK-001",
            "Category": "Geopolitical",
            "Risk_Event": "US IRA FEOC Disqualification",
            "Description": "Violation of FEOC rules due to Chinese extraction/refining, leading to complete loss of the $7,500 tax credit.",
            "Likelihood": 5,
            "Severity": 5,
            "Detectability": 4,
            "Mitigation_Strategy": "Transition fully to FTA-sourced (Canada, Chile, Australia) extraction & non-FEOC refining."
        },
        {
            "Risk_ID": "RSK-002",
            "Category": "Cybersecurity",
            "Risk_Event": "Ransomware Attack on US West Coast Ports",
            "Description": "Cyberattack targeting container terminals at Port of LA/LB, freezing shipping logistics and customs clearance.",
            "Likelihood": 4,
            "Severity": 4,
            "Detectability": 2,
            "Mitigation_Strategy": "Reroute maritime flows to East Coast/Gulf ports (Savannah, Houston) and utilize cross-border North American rail."
        },
        {
            "Risk_ID": "RSK-003",
            "Category": "Climate & Natural Disasters",
            "Risk_Event": "Panama Canal Drought Disruptions",
            "Description": "Drought conditions limit daily vessel transits through Panama Canal, forcing longer routes or intermodal shifts.",
            "Likelihood": 4,
            "Severity": 3,
            "Detectability": 1,
            "Mitigation_Strategy": "Pre-book transit slots or route overland via US West Coast intermodal lanes."
        },
        {
            "Risk_ID": "RSK-004",
            "Category": "Regulatory & Resource Nationalism",
            "Risk_Event": "Indonesian Nickel Tariff / Export Limit",
            "Description": "Indonesian government imposes export tariffs on processed nickel intermediates (MHP) to force domestic cathode production.",
            "Likelihood": 3,
            "Severity": 4,
            "Detectability": 3,
            "Mitigation_Strategy": "Invest in Canadian or Australian mining projects and secure long-term offtake agreements."
        },
        {
            "Risk_ID": "RSK-005",
            "Category": "Macroeconomics",
            "Risk_Event": "Global Freight Rate & Fuel Inflation",
            "Description": "Spike in bunker fuel prices and spot container rates due to regional conflicts and shipping lane detours.",
            "Likelihood": 4,
            "Severity": 3,
            "Detectability": 2,
            "Mitigation_Strategy": "Secure fixed-rate annual ocean freight contracts and shift to regional overland transport."
        }
    ]
    
    df_nodes = pd.DataFrame(nodes)
    df_routes = pd.DataFrame(routes)
    df_risk = pd.DataFrame(risk_library)
    
    excel_path = "data/sc_raw_data.xlsx"
    
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_nodes.to_excel(writer, sheet_name="Nodes", index=False)
        df_routes.to_excel(writer, sheet_name="Routes", index=False)
        df_risk.to_excel(writer, sheet_name="Risk_Library", index=False)
        
    print(f"Data saved to {excel_path} successfully!")
    
    # 4. Apply Excel formatting using openpyxl for visual excellence
    wb = openpyxl.load_workbook(excel_path)
    
    # Stylings
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Navy Blue
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        
        # Header Row styling
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = alignment_center
            cell.border = thin_border
            
        # Data Rows styling
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_body
                cell.border = thin_border
                
                # Format numbers
                val = cell.value
                col_name = ws.cell(row=1, column=col_idx).value
                
                # Format coordinate and cost columns
                if col_name in ["Latitude", "Longitude"]:
                    cell.number_format = '0.0000'
                    cell.alignment = alignment_center
                elif col_name in ["Cost_Per_Ton", "Freight_Cost_Per_Ton"]:
                    cell.number_format = '$#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name in ["Base_Capacity_Tons", "Lead_Time_Days", "Likelihood", "Severity", "Detectability"]:
                    cell.number_format = '#,##0'
                    cell.alignment = alignment_center
                elif col_name in ["Node_ID", "Route_ID", "Risk_ID", "Tier", "FEOC_Status", "Mode", "Scenario"]:
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left
                    
        # Auto-adjust columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(excel_path)
    print("Excel workbook formatted successfully.")

if __name__ == "__main__":
    generate_data()
