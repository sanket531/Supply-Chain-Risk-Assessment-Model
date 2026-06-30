import pandas as pd
import numpy as np
import os
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Set random seed for reproducibility
np.random.seed(42)

def run_simulation():
    print("Loading data from sc_raw_data.xlsx...")
    excel_path = "data/sc_raw_data.xlsx"
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Raw data file not found at {excel_path}. Please run generate_mock_data.py first.")
        
    df_nodes = pd.read_excel(excel_path, sheet_name="Nodes")
    df_routes = pd.read_excel(excel_path, sheet_name="Routes")
    df_risk = pd.read_excel(excel_path, sheet_name="Risk_Library")
    
    # Calculate qualitative RPN
    df_risk["RPN"] = df_risk["Likelihood"] * df_risk["Severity"] * df_risk["Detectability"]
    df_risk = df_risk.sort_values(by="RPN", ascending=False)
    
    print("Running Monte Carlo Simulation (10,000 iterations)...")
    N = 10000
    
    # Sourcing shares for NMC 811 cell material (per ton of cells)
    # Ratios of active materials in NMC 811 cathode/cell:
    # 1 ton of battery cells contains approximately:
    # - 0.10 tons of Lithium (LCE)
    # - 0.15 tons of Nickel
    # - 0.02 tons of Cobalt
    # - 0.12 tons of Graphite (Anode)
    shares = {
        "Lithium": 0.10,
        "Nickel": 0.15,
        "Cobalt": 0.02,
        "Graphite": 0.12
    }
    
    # Safety Stock Buffers (Days)
    safety_stock_asis = 35
    safety_stock_tobe = 15 # Nearshore/domestic sourcing requires less safety stock
    
    # Stockout cost per day of delay
    stockout_cost_per_day = 120000 # Cost of production stoppage
    
    # IRA FEOC Penalty per ton of cell material
    # Loss of $7,500 tax credit per EV pack. Assuming 70 kWh pack, ~0.14 tons of cell material per pack.
    # Penalty translates to ~$53,500 per ton of cell material if FEOC compliant sourcing is failed.
    ira_penalty_per_ton = 53500 
    
    # Pre-allocate results arrays
    results_asis = []
    results_tobe = []
    
    for i in range(N):
        # 1. Roll 2026 Risk Events
        # RSK-002: Port of LA Cyber Ransomware (15% prob)
        la_port_cyber = np.random.rand() < 0.15
        la_delay = np.random.normal(18, 4) if la_port_cyber else 0
        la_surcharge = np.random.normal(600, 100) if la_port_cyber else 0
        
        # RSK-003: Panama Canal Drought (25% prob)
        panama_drought = np.random.rand() < 0.25
        panama_delay = np.random.normal(10, 3) if panama_drought else 0
        panama_surcharge = np.random.normal(300, 50) if panama_drought else 0
        
        # RSK-004: Indonesian Nickel Tariff/Regulation (30% prob)
        idn_nickel_tariff = np.random.rand() < 0.30
        nic_surcharge = np.random.normal(5400, 800) if idn_nickel_tariff else 0
        
        # RSK-005: Freight Inflation (40% prob)
        freight_inflation = np.random.rand() < 0.40
        freight_mult = 1.25 if freight_inflation else 1.0
        
        # General market volatility
        market_freight_var = np.random.normal(1.0, 0.05) # 5% SD
        market_material_var = np.random.normal(1.0, 0.02) # 2% SD
        
        # -----------------------------------------------------------------
        # AS-IS SCENARIO SIMULATION (Vulnerable, Global Sourcing)
        # -----------------------------------------------------------------
        # Sourcing paths for As-Is:
        # Lithium: 50% Chile (MIN_LITH_CHL -> REF_LITH_CHL -> CAM_KOR), 50% Australia (MIN_LITH_AUS -> CAM_KOR)
        # Cobalt: DRC (MIN_COB_DRC -> REF_COB_CHN -> CAM_KOR)
        # Nickel: Indonesia (MIN_NIC_IDN -> REF_NIC_CHN -> CAM_KOR)
        # Graphite: China (MIN_GRA_CHN -> REF_GRA_CHN -> CAM_KOR)
        # CAM to Gigafactory: CAM_KOR -> GIGA_USA (via LA Port)
        
        # Lead Times (Base + Volatility + Risk delays)
        lt_var = lambda: np.random.normal(0, 1.5)
        
        # Lithium Lead Time
        lt_lith_chile = 1 + 28 + lt_var() + (panama_delay if panama_drought else 0)
        lt_lith_aus = 18 + lt_var()
        lt_lith_asis = 0.5 * lt_lith_chile + 0.5 * lt_lith_aus
        
        # Cobalt Lead Time
        lt_cob_asis = 32 + 3 + lt_var() + (panama_delay if panama_drought else 0)
        
        # Nickel Lead Time
        lt_nic_asis = 12 + 3 + lt_var()
        
        # Graphite Lead Time
        lt_gra_asis = 4 + 4 + lt_var()
        
        # CAM to Assembly Lead Time
        lt_cam_asis = 25 + lt_var() + (la_delay if la_port_cyber else 0)
        
        # Max component lead time to CAM plant + CAM to GIGA
        max_lead_time_asis = max(lt_lith_asis, lt_cob_asis, lt_nic_asis, lt_gra_asis) + lt_cam_asis
        
        # Costs (Base + Volatility + Surcharges)
        # Material Costs (Mining + Refining)
        cost_lith_asis = (shares["Lithium"] * ((8000 + 2000) * market_material_var))
        cost_cob_asis = (shares["Cobalt"] * ((25000 + 4000) * market_material_var))
        cost_nic_asis = (shares["Nickel"] * ((12000 + 3000) * market_material_var + nic_surcharge))
        cost_gra_asis = (shares["Graphite"] * ((3000 + 1500) * market_material_var))
        
        # Freight Costs
        f_lith_chile = (200 + 1200) * freight_mult * market_freight_var + (panama_surcharge if panama_drought else 0)
        f_lith_aus = 900 * freight_mult * market_freight_var
        f_lith_asis = shares["Lithium"] * (0.5 * f_lith_chile + 0.5 * f_lith_aus)
        
        f_cob_asis = shares["Cobalt"] * ((1500 + 300) * freight_mult * market_freight_var + (panama_surcharge if panama_drought else 0))
        f_nic_asis = shares["Nickel"] * ((600 + 300) * freight_mult * market_freight_var)
        f_gra_asis = shares["Graphite"] * ((400 + 350) * freight_mult * market_freight_var)
        
        # CAM to Assembly Freight (LA Port route)
        f_cam_asis = 1800 * freight_mult * market_freight_var + (la_surcharge if la_port_cyber else 0)
        
        # Manufacturing value add (CAM synthesis + Cell assembly)
        mfg_asis = (15000 * shares["Lithium"] + 10000) # Base assembly + CAM synthesis cost
        
        total_baseline_cost_asis = (cost_lith_asis + cost_cob_asis + cost_nic_asis + cost_gra_asis +
                                    f_lith_asis + f_cob_asis + f_nic_asis + f_gra_asis + f_cam_asis + mfg_asis)
        
        # Risk Penalties
        ira_penalty_asis = ira_penalty_per_ton # Always fails FEOC rules
        
        # Stockout Cost
        stockout_days_asis = max(0, max_lead_time_asis - safety_stock_asis)
        stockout_cost_asis = stockout_days_asis * stockout_cost_per_day / 100 # Distributed cost per ton
        
        total_tcor_asis = total_baseline_cost_asis + ira_penalty_asis + stockout_cost_asis
        
        results_asis.append({
            "Trial": i,
            "Lithium_Lead_Time": lt_lith_asis,
            "Cobalt_Lead_Time": lt_cob_asis,
            "Nickel_Lead_Time": lt_nic_asis,
            "Graphite_Lead_Time": lt_gra_asis,
            "Total_Lead_Time": max_lead_time_asis,
            "Material_Cost": cost_lith_asis + cost_cob_asis + cost_nic_asis + cost_gra_asis,
            "Freight_Cost": f_lith_asis + f_cob_asis + f_nic_asis + f_gra_asis + f_cam_asis,
            "Mfg_Cost": mfg_asis,
            "IRA_Penalty": ira_penalty_asis,
            "Stockout_Cost": stockout_cost_asis,
            "Total_Cost_TCOR": total_tcor_asis,
            "Disruption_LA_Port": int(la_port_cyber),
            "Disruption_Panama": int(panama_drought),
            "Disruption_Nickel_Tariff": int(idn_nickel_tariff)
        })
        
        # -----------------------------------------------------------------
        # TO-BE SCENARIO SIMULATION (Resilient, FTA-Compliant, Nearshore)
        # -----------------------------------------------------------------
        # Sourcing paths for To-Be:
        # Lithium: 50% Chile (REF_LITH_CHL -> CAM_USA), 50% Canada (MIN_LITH_CAN -> CAM_USA)
        # Cobalt: Canada to Finland (MIN_COB_CAN -> REF_COB_FIN -> CAM_USA)
        # Nickel: Canada to Canada (MIN_NIC_CAN -> REF_NIC_CAN -> CAM_USA)
        # Graphite: US synthetic (MIN_GRA_USA -> REF_GRA_USA -> CAM_USA)
        # CAM to Gigafactory: CAM_USA -> GIGA_USA (Truck/Rail)
        
        # Lithium Lead Time
        lt_lith_chile_tb = 1 + 22 + lt_var() # Direct to East Coast
        lt_lith_can_tb = 6 + lt_var()
        lt_lith_tobe = 0.5 * lt_lith_chile_tb + 0.5 * lt_lith_can_tb
        
        # Cobalt Lead Time
        lt_cob_tobe = 16 + 14 + lt_var()
        
        # Nickel Lead Time
        lt_nic_tobe = 5 + 4 + lt_var()
        
        # Graphite Lead Time
        lt_gra_tobe = 2 + 2 + lt_var()
        
        # CAM to Assembly Lead Time
        lt_cam_tobe = 2 + lt_var()
        
        max_lead_time_tobe = max(lt_lith_tobe, lt_cob_tobe, lt_nic_tobe, lt_gra_tobe) + lt_cam_tobe
        
        # Costs (Base + Volatility + Surcharges)
        # Material Costs
        cost_lith_tobe = shares["Lithium"] * (0.5 * (8000 + 2000) + 0.5 * 11000) * market_material_var
        cost_cob_tobe = shares["Cobalt"] * (32000 + 6000) * market_material_var
        cost_nic_tobe = shares["Nickel"] * (15000 + 4500) * market_material_var # Canadian Nickel is FEOC-free, no surcharge
        cost_gra_tobe = shares["Graphite"] * (7500 + 3000) * market_material_var
        
        # Freight Costs
        f_lith_chile_tb = (200 + 1400) * freight_mult * market_freight_var # Direct to East Coast, bypass Panama drought / LA Port
        f_lith_can_tb = 800 * market_freight_var
        f_lith_tobe = shares["Lithium"] * (0.5 * f_lith_chile_tb + 0.5 * f_lith_can_tb)
        
        f_cob_tobe = shares["Cobalt"] * (1100 + 1300) * freight_mult * market_freight_var
        f_nic_tobe = shares["Nickel"] * (600 + 500) * market_freight_var
        f_gra_tobe = shares["Graphite"] * (300 + 250) * market_freight_var
        
        # CAM to Assembly Freight (Domestic truck)
        f_cam_tobe = 400 * market_freight_var
        
        mfg_tobe = (19000 * shares["Lithium"] + 10000) # Domestic CAM synthesis + Cell assembly
        
        total_baseline_cost_tobe = (cost_lith_tobe + cost_cob_tobe + cost_nic_tobe + cost_gra_tobe +
                                    f_lith_tobe + f_cob_tobe + f_nic_tobe + f_gra_tobe + f_cam_tobe + mfg_tobe)
        
        # Risk Penalties
        ira_penalty_tobe = 0 # 100% FEOC free, qualifies for tax credit
        
        # Stockout Cost
        stockout_days_tobe = max(0, max_lead_time_tobe - safety_stock_tobe)
        stockout_cost_tobe = stockout_days_tobe * stockout_cost_per_day / 100 # Distributed cost per ton
        
        total_tcor_tobe = total_baseline_cost_tobe + ira_penalty_tobe + stockout_cost_tobe
        
        results_tobe.append({
            "Trial": i,
            "Lithium_Lead_Time": lt_lith_tobe,
            "Cobalt_Lead_Time": lt_cob_tobe,
            "Nickel_Lead_Time": lt_nic_tobe,
            "Graphite_Lead_Time": lt_gra_tobe,
            "Total_Lead_Time": max_lead_time_tobe,
            "Material_Cost": cost_lith_tobe + cost_cob_tobe + cost_nic_tobe + cost_gra_tobe,
            "Freight_Cost": f_lith_tobe + f_cob_tobe + f_nic_tobe + f_gra_tobe + f_cam_tobe,
            "Mfg_Cost": mfg_tobe,
            "IRA_Penalty": ira_penalty_tobe,
            "Stockout_Cost": stockout_cost_tobe,
            "Total_Cost_TCOR": total_tcor_tobe,
            "Disruption_LA_Port": int(la_port_cyber),
            "Disruption_Panama": int(panama_drought),
            "Disruption_Nickel_Tariff": int(idn_nickel_tariff)
        })

    df_asis_runs = pd.DataFrame(results_asis)
    df_tobe_runs = pd.DataFrame(results_tobe)
    
    # 2. Compute Summary Statistics
    summary_stats = []
    for sc_name, df in [("As-Is (Unmitigated)", df_asis_runs), ("To-Be (Resilient)", df_tobe_runs)]:
        summary_stats.append({
            "Scenario": sc_name,
            "Mean_Total_Lead_Time_Days": df["Total_Lead_Time"].mean(),
            "Max_Total_Lead_Time_Days": df["Total_Lead_Time"].max(),
            "Lead_Time_90th_Percentile": df["Total_Lead_Time"].quantile(0.90),
            "Lead_Time_99th_Percentile": df["Total_Lead_Time"].quantile(0.99),
            "Mean_Material_Cost": df["Material_Cost"].mean(),
            "Mean_Freight_Cost": df["Freight_Cost"].mean(),
            "Mean_Mfg_Cost": df["Mfg_Cost"].mean(),
            "Mean_IRA_Penalty": df["IRA_Penalty"].mean(),
            "Mean_Stockout_Cost": df["Stockout_Cost"].mean(),
            "Mean_Total_Cost_TCOR": df["Total_Cost_TCOR"].mean(),
            "Cost_90th_Percentile_VaR": df["Total_Cost_TCOR"].quantile(0.90),
            "Cost_99th_Percentile_VaR": df["Total_Cost_TCOR"].quantile(0.99),
            "Stockout_Probability": (df["Stockout_Cost"] > 0).mean()
        })
    df_summary = pd.DataFrame(summary_stats)
    
    # Export CSVs for Tableau
    print("Exporting Tableau CSVs...")
    df_summary.to_csv("data/tableau_summary.csv", index=False)
    
    # Export 1,000 runs to keep file size reasonable for Tableau dashboard
    df_asis_runs_sample = df_asis_runs.sample(1000, random_state=42).copy()
    df_asis_runs_sample["Scenario"] = "As-Is (Unmitigated)"
    df_tobe_runs_sample = df_tobe_runs.sample(1000, random_state=42).copy()
    df_tobe_runs_sample["Scenario"] = "To-Be (Resilient)"
    
    df_tableau_runs = pd.concat([df_asis_runs_sample, df_tobe_runs_sample])
    df_tableau_runs.to_csv("data/tableau_simulation_runs.csv", index=False)
    
    # Export nodes and routes mapped to simulated parameters
    df_nodes_tbl = df_nodes.copy()
    # Add simulated risk scores
    df_nodes_tbl["Simulated_Avg_Cost"] = df_nodes_tbl["Cost_Per_Ton"]
    df_nodes_tbl["Vulnerability_Index"] = df_nodes_tbl.apply(
        lambda r: 5 if r["FEOC_Status"] == "Yes" else (3 if r["Country"] == "DRC" else 1), axis=1
    )
    df_nodes_tbl.to_csv("data/tableau_nodes.csv", index=False)
    
    df_routes_tbl = df_routes.copy()
    # Add average lead times
    df_routes_tbl.to_csv("data/tableau_routes.csv", index=False)
    
    # 3. Create Plots
    print("Generating Cost and Lead Time Distribution plots...")
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    
    # Cost Plot
    axes[0].hist(df_asis_runs["Total_Cost_TCOR"], bins=50, alpha=0.6, label="As-Is (Unmitigated)", color="#d9534f")
    axes[0].hist(df_tobe_runs["Total_Cost_TCOR"], bins=50, alpha=0.6, label="To-Be (Resilient)", color="#5cb85c")
    axes[0].axvline(df_asis_runs["Total_Cost_TCOR"].mean(), color="#d9534f", linestyle="dashed", linewidth=1.5, label="As-Is Mean")
    axes[0].axvline(df_tobe_runs["Total_Cost_TCOR"].mean(), color="#5cb85c", linestyle="dashed", linewidth=1.5, label="To-Be Mean")
    axes[0].set_title("Total Cost of Risk (TCOR) Distribution ($/Ton)", fontsize=12, fontweight='bold')
    axes[0].set_xlabel("Total Cost of Risk ($/Ton)")
    axes[0].set_ylabel("Frequency")
    axes[0].grid(True, linestyle=":", alpha=0.6)
    axes[0].legend()
    
    # Lead Time Plot
    axes[1].hist(df_asis_runs["Total_Lead_Time"], bins=50, alpha=0.6, label="As-Is (Unmitigated)", color="#d9534f")
    axes[1].hist(df_tobe_runs["Total_Lead_Time"], bins=50, alpha=0.6, label="To-Be (Resilient)", color="#5cb85c")
    axes[1].axvline(df_asis_runs["Total_Lead_Time"].mean(), color="#d9534f", linestyle="dashed", linewidth=1.5, label="As-Is Mean")
    axes[1].axvline(df_tobe_runs["Total_Lead_Time"].mean(), color="#5cb85c", linestyle="dashed", linewidth=1.5, label="To-Be Mean")
    axes[1].set_title("Total Supply Chain Lead Time (Days)", fontsize=12, fontweight='bold')
    axes[1].set_xlabel("Lead Time (Days)")
    axes[1].set_ylabel("Frequency")
    axes[1].grid(True, linestyle=":", alpha=0.6)
    axes[1].legend()
    
    plt.tight_layout()
    plt.savefig("visualizations/simulation_results.png", dpi=300)
    plt.close()
    
    # 4. Generate Network Risk Concentration Map
    print("Generating Network Risk Concentration Map...")
    plt.figure(figsize=(12, 7))
    
    # Plot routes as lines
    # Set coordinates for visual mapping
    for _, route in df_routes.iterrows():
        orig_node = df_nodes[df_nodes["Node_ID"] == route["Origin"]].iloc[0]
        dest_node = df_nodes[df_nodes["Node_ID"] == route["Destination"]].iloc[0]
        
        color = "#e74c3c" if route["Scenario"] == "As-Is" else "#2ecc71"
        style = ":" if route["Scenario"] == "As-Is" else "-"
        
        plt.plot([orig_node["Longitude"], dest_node["Longitude"]], 
                 [orig_node["Latitude"], dest_node["Latitude"]], 
                 color=color, linestyle=style, alpha=0.4, linewidth=1.5)
                 
    # Plot nodes
    for _, node in df_nodes.iterrows():
        # Size represents capacity, color represents risk exposure
        size = node["Base_Capacity_Tons"] / 30
        
        if node["FEOC_Status"] == "Yes":
            color = "#c0392b" # Red (High Risk / FEOC)
        elif node["Country"] in ["DRC"]:
            color = "#f39c12" # Orange (Medium Geopolitical Risk)
        elif node["Tier"].startswith("Tier-1"):
            color = "#2980b9" # Blue (Assembly Node)
        else:
            color = "#27ae60" # Green (Low Risk / Compliant)
            
        plt.scatter(node["Longitude"], node["Latitude"], s=size, color=color, alpha=0.8, edgecolors="black", linewidth=0.5, zorder=5)
        # Add labels for key logistics hubs
        if node["Node_ID"] in ["GIGA_USA", "CAM_KOR", "CAM_USA", "REF_GRA_CHN", "REF_NIC_CHN", "MIN_COB_DRC", "REF_LITH_CHL"]:
            plt.text(node["Longitude"] + 2, node["Latitude"] + 2, node["Name"], fontsize=8, fontweight="bold", alpha=0.7)
            
    plt.title("2026 EV Battery Supply Chain: Network Risk Concentration Map", fontsize=14, fontweight="bold")
    plt.xlabel("Longitude")
    plt.ylabel("Latitude")
    plt.grid(True, linestyle="--", alpha=0.3)
    
    # Legend
    high_risk_patch = plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#c0392b', label='High Risk (FEOC Node)', markersize=10)
    med_risk_patch = plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#f39c12', label='Medium Risk (DRC Node)', markersize=10)
    low_risk_patch = plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#27ae60', label='Low Risk / Compliant Node', markersize=10)
    assembly_patch = plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#2980b9', label='Assembly Node', markersize=10)
    
    asis_line = plt.Line2D([0], [0], color='#e74c3c', linestyle=":", label='As-Is Route (Unmitigated)', linewidth=2)
    tobe_line = plt.Line2D([0], [0], color='#2ecc71', linestyle="-", label='To-Be Route (Resilient)', linewidth=2)
    
    plt.legend(handles=[high_risk_patch, med_risk_patch, low_risk_patch, assembly_patch, asis_line, tobe_line], loc="lower left")
    
    plt.tight_layout()
    plt.savefig("visualizations/network_risk_map.png", dpi=300)
    plt.close()
    
    # 5. Output beautiful Excel sheets
    print("Writing results back to Excel...")
    wb = openpyxl.load_workbook(excel_path)
    
    # Update Risk Library sheet with RPN formulas
    ws_risk = wb["Risk_Library"]
    ws_risk.cell(row=1, column=ws_risk.max_column + 1, value="RPN").font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    # Apply RPN Formula in Excel
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Re-write column headers and formula
    for row_idx in range(2, ws_risk.max_row + 1):
        cell_formula = f"=E{row_idx}*F{row_idx}*G{row_idx}" # Likelihood * Severity * Detectability
        rpn_cell = ws_risk.cell(row=row_idx, column=8, value=cell_formula)
        rpn_cell.font = Font(name="Segoe UI", size=10)
        rpn_cell.border = thin_border
        rpn_cell.alignment = Alignment(horizontal="center")
        rpn_cell.number_format = '#,##0'
        
    # Write Risk Scorecard Summary sheet
    if "Risk_Scorecard" in wb.sheetnames:
        del wb["Risk_Scorecard"]
    ws_scorecard = wb.create_sheet(title="Risk_Scorecard")
    ws_scorecard.views.sheetView[0].showGridLines = True
    
    # Sheet Title
    ws_scorecard.cell(row=1, column=1, value="2026 SUPPLY CHAIN RISK SCORECARD & SCENARIO COMPARISON").font = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
    
    # Add summary table
    headers = [
        "Metric", "As-Is (Unmitigated)", "To-Be (Resilient)", "Delta / Savings", "% Improvement"
    ]
    for col_idx, h in enumerate(headers, 1):
        c = ws_scorecard.cell(row=3, column=col_idx, value=h)
        c.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
        
    metrics = [
        ("Mean Lead Time (Days)", df_summary.loc[0, "Mean_Total_Lead_Time_Days"], df_summary.loc[1, "Mean_Total_Lead_Time_Days"], "Days"),
        ("90th Percentile Lead Time (Days)", df_summary.loc[0, "Lead_Time_90th_Percentile"], df_summary.loc[1, "Lead_Time_90th_Percentile"], "Days"),
        ("Stockout Probability (%)", df_summary.loc[0, "Stockout_Probability"] * 100, df_summary.loc[1, "Stockout_Probability"] * 100, "Pct"),
        ("Mean Material Sourcing Cost ($/Ton)", df_summary.loc[0, "Mean_Material_Cost"], df_summary.loc[1, "Mean_Material_Cost"], "USD"),
        ("Mean Freight Logistics Cost ($/Ton)", df_summary.loc[0, "Mean_Freight_Cost"], df_summary.loc[1, "Mean_Freight_Cost"], "USD"),
        ("Mean Manufacturing & Assembly ($/Ton)", df_summary.loc[0, "Mean_Mfg_Cost"], df_summary.loc[1, "Mean_Mfg_Cost"], "USD"),
        ("Mean IRA Geopolitical Penalty ($/Ton)", df_summary.loc[0, "Mean_IRA_Penalty"], df_summary.loc[1, "Mean_IRA_Penalty"], "USD"),
        ("Mean Expected Stockout Cost ($/Ton)", df_summary.loc[0, "Mean_Stockout_Cost"], df_summary.loc[1, "Mean_Stockout_Cost"], "USD"),
        ("Total Cost of Risk - TCOR ($/Ton)", df_summary.loc[0, "Mean_Total_Cost_TCOR"], df_summary.loc[1, "Mean_Total_Cost_TCOR"], "USD"),
        ("90th Percentile Cost VaR ($/Ton)", df_summary.loc[0, "Cost_90th_Percentile_VaR"], df_summary.loc[1, "Cost_90th_Percentile_VaR"], "USD"),
    ]
    
    font_body = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light Blue
    
    for idx, (m_name, asis_val, tobe_val, m_type) in enumerate(metrics, 4):
        ws_scorecard.cell(row=idx, column=1, value=m_name).font = font_bold if "Total" in m_name or "TCOR" in m_name else font_body
        ws_scorecard.cell(row=idx, column=2, value=asis_val).font = font_bold if "Total" in m_name or "TCOR" in m_name else font_body
        ws_scorecard.cell(row=idx, column=3, value=tobe_val).font = font_bold if "Total" in m_name or "TCOR" in m_name else font_body
        
        # Formulas for Delta and % Improvement
        delta_formula = f"=B{idx}-C{idx}"
        pct_formula = f"=D{idx}/B{idx}"
        
        c_delta = ws_scorecard.cell(row=idx, column=4, value=delta_formula)
        c_pct = ws_scorecard.cell(row=idx, column=5, value=pct_formula)
        
        c_delta.font = font_bold if "Total" in m_name or "TCOR" in m_name else font_body
        c_pct.font = font_bold if "Total" in m_name or "TCOR" in m_name else font_body
        
        # Border & Fills
        for col_idx in range(1, 6):
            cell = ws_scorecard.cell(row=idx, column=col_idx)
            cell.border = thin_border
            if "Total" in m_name or "TCOR" in m_name:
                cell.fill = fill_total
                
            # Number formats
            if col_idx in [2, 3, 4]:
                if m_type == "USD":
                    cell.number_format = '$#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif m_type == "Days":
                    cell.number_format = '0.0'
                    cell.alignment = Alignment(horizontal="center")
                elif m_type == "Pct":
                    cell.number_format = '0.0'
                    cell.alignment = Alignment(horizontal="center")
            elif col_idx == 5:
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
                
    # Auto-adjust column widths
    for col in ws_scorecard.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_scorecard.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(excel_path)
    # Also save a final copy as the named deliverable
    wb.save("sc_risk_assessment.xlsx")
    print(f"Results saved to {excel_path} and sc_risk_assessment.xlsx successfully!")

if __name__ == "__main__":
    run_simulation()
